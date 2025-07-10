import openpyxl
from datetime import datetime

def format_date(cell):
    if isinstance(cell, datetime):
        return cell.strftime("%d.%m.%Y")
    elif isinstance(cell, str):
        return cell.strip()
    return "Не указано"

def load_employees(path="employees.xlsx"):
    wb = openpyxl.load_workbook(path)
    employees = []

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not row or not row[1]:
                continue

            employees.append({
                "name": str(row[1]).strip(),
                "position": row[2],
                "birth": format_date(row[3]),
                "start_date": format_date(row[4]),
                "email": str(row[5]) if row[5] else "—",
                "phone": str(row[6]) if row[6] else "—",
                "contract_type": row[7] or "—",
                "sheet": sheet_name,
            })
    return employees
